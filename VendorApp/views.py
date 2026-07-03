from django.shortcuts import render, redirect
from django.contrib import messages

from django.http import JsonResponse, HttpResponse
import csv

from django.views.decorators.http import require_http_methods

from GuestApp.models import tbl_login, tbl_vendor

from AdminApp.models import tbl_district, tbl_location, tbl_categorys, tbl_subcategory

from VendorApp.models import tbl_product, tbl_furniture_details, tbl_appliance_details, tbl_product_image

from CustomerApp.models import tbl_order, tbl_order_item, tbl_return_request

from decimal import Decimal

from datetime import datetime
import smtplib
from email.message import EmailMessage
from django.views.decorators.cache import cache_control


def send_email_notification(subject, body, to_email, html_body=None):
    """
    Helper function to send email notifications.
    """
    try:
        msg = EmailMessage()
        msg.set_content(body)
        if html_body:
            msg.add_alternative(html_body, subtype='html')
            
        msg['Subject'] = subject
        msg['From'] = "urbanrentalsofficial@gmail.com"
        msg['To'] = to_email
        
        # Connect to Gmail's SMTP server
        server = smtplib.SMTP_SSL("smtp.gmail.com", 465)
        server.login("urbanrentalsofficial@gmail.com", "kilw hwzc jdkz xlmp")
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print(f"Error sending email: {e}")
        return False



# Create your views here.

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def vendor_home(request):
    
    if 'vendor' not in request.session:
        return HttpResponse("<script>alert('Authentication Required! Please login.');window.location='/guestapp/login/';</script>")
    
    try:
        vendor = tbl_vendor.objects.get(login_id=request.session['vendor'])
        
        # Get dashboard statistics
        total_products = tbl_product.objects.filter(vendor=vendor).count()
        
        # Get active orders (filter by vendor's products)
        vendor_orders = tbl_order.objects.filter(
            items__product__vendor=vendor,
            status__in=['pending', 'confirmed', 'paid']
        ).distinct().count()
        
        # Get pending return requests
        pending_returns = tbl_return_request.objects.filter(
            order_item__product__vendor=vendor,
            status='requested'
        ).distinct().count()
        
        # Calculate total revenue from active/completed orders
        from django.db.models import Sum, Count, F, DecimalField
        from django.db.models.functions import Coalesce, TruncMonth, TruncYear
        import json
        
        # Consider revenue from Paid and Completed orders
        # Including 'confirmed' to show potential revenue if desired, or just paid/completed.
        # Let's include 'confirmed' as well if the payment flow is not strictly enforced yet.
        revenue_statuses = ['paid', 'completed', 'rented', 'returned', 'confirmed']
        
        revenue_orders = tbl_order.objects.filter(
            items__product__vendor=vendor,
            status__in=revenue_statuses 
        ).distinct()
        
        total_revenue = Decimal('0')
        # Efficient calculation using aggregation instead of looping
        revenue_aggregate = tbl_order_item.objects.filter(
            product__vendor=vendor,
            order__status__in=revenue_statuses
        ).aggregate(total=Sum('grand_total'))
        
        total_revenue = revenue_aggregate['total'] if revenue_aggregate['total'] else Decimal('0')

        # Chart Data: Monthly Sales for Current Year
        # Filter for current year and group by month
        from datetime import datetime
        current_year = datetime.now().year
        
        # Get sales for the current year
        monthly_sales = tbl_order_item.objects.filter(
            product__vendor=vendor,
            order__status__in=revenue_statuses,
            order__created_at__year=current_year
        ).annotate(month=TruncMonth('order__created_at'))\
         .values('month')\
         .annotate(total_revenue=Sum('grand_total'))\
         .order_by('month')

        # Convert QuerySet to list of dictionaries
        monthly_sales_list = list(monthly_sales)
        
        # Create a dictionary for quick lookup: {month_number: revenue}
        sales_data_dict = {}
        for entry in monthly_sales_list:
            if entry['month']:
                sales_data_dict[entry['month'].month] = float(entry['total_revenue'])
        
        # Ensure all 12 months are represented (Jan-Dec)
        sales_dates = []
        sales_amounts = []
        import calendar
        
        for i in range(1, 13):
            month_name = calendar.month_name[i]
            sales_dates.append(month_name)
            sales_amounts.append(sales_data_dict.get(i, 0.0))

        # Chart Data: Product Popularity (Top 5)
        top_products = tbl_order_item.objects.filter(product__vendor=vendor)\
            .values('product__product_name')\
            .annotate(rent_count=Count('product'))\
            .order_by('-rent_count')[:5]
        
        product_labels = [entry['product__product_name'] for entry in top_products]
        product_data = [entry['rent_count'] for entry in top_products]
        
        context = {
            'vendor': vendor,
            'total_products': total_products,
            'total_orders': vendor_orders,
            'pending_returns': pending_returns,
            'total_revenue': total_revenue,
            'sales_dates': sales_dates,
            'sales_amounts': sales_amounts,
            'current_year': current_year,
            'product_labels': product_labels,
            'product_data': product_data,
        }
        return render(request, 'Vendor/index.html', context)
    except tbl_vendor.DoesNotExist:
        messages.error(request, "Vendor account not found")
        return redirect('login')



def my_profile(request):

    """Vendor Profile View"""

    try:

        # Get vendor ID from session

        vendor_login_id = request.session.get('vendor')

        if not vendor_login_id:

            return redirect('login')

            

        vendor = tbl_vendor.objects.get(login_id=vendor_login_id)

        

        context = {

            'vendor': vendor

        }

        return render(request, 'Vendor/my_profile.html', context)

    except tbl_vendor.DoesNotExist:

        messages.error(request, "Vendor account not found")

        return redirect('vendor_home')









# ======================== PRODUCT MANAGEMENT ========================



def add_product(request):

    """Add new product for rent"""

    try:

        vendor = tbl_vendor.objects.get(login_id=request.session.get('vendor'))

    except tbl_vendor.DoesNotExist:

        messages.error(request, "Vendor account not found")

        return redirect('vendor_home')



    if request.method == 'POST':

        # Basic Product Fields

        category_id = request.POST.get('category')

        subcategory_id = request.POST.get('subcategory')

        product_name = request.POST.get('product_name')

        product_description = request.POST.get('product_description')

        brand = request.POST.get('brand')

        rent_price_per_day = request.POST.get('rent_price_per_day')

        rent_price_per_month = request.POST.get('rent_price_per_month')

        security_deposit = request.POST.get('security_deposit')

        stock_quantity = request.POST.get('stock_quantity')

        product_image = request.FILES.get('product_image')

        

        # Furniture Details

        furniture_type = request.POST.get('furniture_type')

        furniture_size = request.POST.get('furniture_size')

        furniture_thickness = request.POST.get('furniture_thickness')

        furniture_material = request.POST.get('furniture_material')

        furniture_features = request.POST.get('furniture_features')

        

        # Appliance Details

        appliance_type = request.POST.get('appliance_type')

        appliance_power_rating = request.POST.get('appliance_power_rating')

        appliance_voltage = request.POST.get('appliance_voltage')

        appliance_warranty = request.POST.get('appliance_warranty')



        # Validation

        errors = []

        if not category_id:

            errors.append("Category is required")

        if not subcategory_id:

            errors.append("Subcategory is required")

        if not product_name or len(product_name.strip()) == 0:

            errors.append("Product name is required")

        if not product_description or len(product_description.strip()) == 0:

            errors.append("Description is required")

        if not product_image:

            errors.append("Product image is required")

        

        # At least one pricing option

        if not rent_price_per_day and not rent_price_per_month:

            errors.append("Enter at least one rental price (daily or monthly)")



        # Validate prices

        try:

            if rent_price_per_day:

                rent_price_per_day = float(rent_price_per_day)

                if rent_price_per_day < 0:

                    errors.append("Daily price cannot be negative")

        except ValueError:

            errors.append("Daily price must be a valid number")



        try:

            if rent_price_per_month:

                rent_price_per_month = float(rent_price_per_month)

                if rent_price_per_month < 0:

                    errors.append("Monthly price cannot be negative")

        except ValueError:

            errors.append("Monthly price must be a valid number")



        # Validate security deposit

        try:

            if security_deposit:

                security_deposit = float(security_deposit)

                if security_deposit < 0:

                    errors.append("Security deposit cannot be negative")

            else:

                security_deposit = 0

        except ValueError:

            errors.append("Security deposit must be a valid number")



        # Validate quantity

        try:

            stock_quantity = int(stock_quantity) if stock_quantity else 1

            if stock_quantity <= 0:

                errors.append("Quantity must be at least 1")

        except ValueError:

            errors.append("Quantity must be a valid number")



        # Validate image

        if product_image:

            allowed_types = ['image/jpeg', 'image/png', 'image/webp']

            if product_image.content_type not in allowed_types:

                errors.append("Product image must be JPEG, PNG, or WebP")



        # Validate category-specific details

        try:

            category = tbl_categorys.objects.get(category_id=category_id)

            # Check if furniture details are required (for Furniture category)

            if category.category_name.lower() == 'furniture':

                if not furniture_type:

                    errors.append("Furniture type is required")

                if not furniture_size:

                    errors.append("Furniture size is required")

                if not furniture_material:

                    errors.append("Furniture material is required")

            

            # Check if appliance details are required (for Appliances category)

            elif category.category_name.lower() == 'appliances':

                if not appliance_type:

                    errors.append("Appliance type is required")

                if not appliance_power_rating:

                    errors.append("Power rating is required")

                if not appliance_voltage:

                    errors.append("Voltage is required")

        except tbl_categorys.DoesNotExist:

            errors.append("Selected category not found")



        if errors:

            categories = tbl_categorys.objects.all()

            for error in errors:

                messages.error(request, error)

            return render(request, 'Vendor/add_product.html', {

                'categories': categories,

                'form_data': request.POST

            })



        try:

            category = tbl_categorys.objects.get(category_id=category_id)

            subcategory = tbl_subcategory.objects.get(subcategory_id=subcategory_id)



            # Verify subcategory belongs to selected category

            if subcategory.category_id != int(category_id):

                messages.error(request, "Invalid subcategory selection")

                return redirect('add_product')



            # Create product

            product = tbl_product.objects.create(

                vendor=vendor,

                category=category,

                subcategory=subcategory,

                product_name=product_name.strip(),

                product_description=product_description.strip(),

                brand=brand.strip() if brand else None,

                rent_price_per_day=rent_price_per_day if rent_price_per_day else None,

                rent_price_per_month=rent_price_per_month if rent_price_per_month else None,

                security_deposit=security_deposit,

                stock_quantity=stock_quantity,

                product_image=product_image,

                is_available=True,

                status='Active'

            )



            # Create furniture details if category is Furniture

            if category.category_name.lower() == 'furniture':

                tbl_furniture_details.objects.create(

                    product=product,

                    furniture_type=furniture_type.strip(),

                    size=furniture_size.strip(),

                    thickness=furniture_thickness.strip() if furniture_thickness else '',

                    material=furniture_material.strip(),

                    extra_features=furniture_features.strip() if furniture_features else None

                )



            # Create appliance details if category is Appliances

            elif category.category_name.lower() == 'appliances':

                tbl_appliance_details.objects.create(

                    product=product,

                    operation_type=appliance_type.strip(),

                    power_rating=appliance_power_rating.strip(),

                    voltage=appliance_voltage.strip(),

                    warranty_period=appliance_warranty.strip() if appliance_warranty else ''

                )



            # Handle additional product images if provided

            additional_images = request.FILES.getlist('additional_images')

            for image in additional_images:

                if image:

                    tbl_product_image.objects.create(

                        product=product,

                        image=image

                    )



            messages.success(request, f"âœ… Product '{product_name}' added successfully!")

            return redirect('vendor_products')



        except tbl_categorys.DoesNotExist:

            messages.error(request, "Selected category not found")

            return redirect('add_product')

        except tbl_subcategory.DoesNotExist:

            messages.error(request, "Selected subcategory not found")

            return redirect('add_product')

        except Exception as e:

            import traceback

            error_details = traceback.format_exc()

            print(f"Error adding product: {error_details}")  # Print to console for debugging

            messages.error(request, f"Error adding product: {str(e)}")

            return redirect('add_product')



    categories = tbl_categorys.objects.all()

    return render(request, 'Vendor/add_product.html', {

        'categories': categories

    })





def vendor_products(request):

    """Display vendor's products"""
    
    if 'vendor' not in request.session:
        messages.warning(request, "Please login to view your products")
        return redirect('login')

    try:

        vendor = tbl_vendor.objects.get(login_id=request.session.get('vendor'))

        products = tbl_product.objects.filter(vendor=vendor)

        return render(request, 'Vendor/products.html', {

            'products': products,

            'vendor': vendor

        })

    except tbl_vendor.DoesNotExist:

        messages.error(request, "Vendor account not found")

        return redirect('login')





def edit_product(request, product_id):

    """Edit an existing product"""

    try:

        vendor = tbl_vendor.objects.get(login_id=request.session.get('vendor'))

        product = tbl_product.objects.get(product_id=product_id, vendor=vendor)

    except tbl_vendor.DoesNotExist:

        messages.error(request, "Vendor account not found")

        return redirect('vendor_home')

    except tbl_product.DoesNotExist:

        messages.error(request, "Product not found or you don't have permission to edit it")

        return redirect('vendor_products')



    if request.method == 'POST':

        # Get form data

        product_name = request.POST.get('product_name')

        product_description = request.POST.get('product_description')

        brand = request.POST.get('brand')

        rent_price_per_day = request.POST.get('rent_price_per_day')

        rent_price_per_month = request.POST.get('rent_price_per_month')

        security_deposit = request.POST.get('security_deposit')

        stock_quantity = request.POST.get('stock_quantity')

        product_image = request.FILES.get('product_image')

        is_available = request.POST.get('is_available') == 'on'

        

        # Furniture Details

        furniture_type = request.POST.get('furniture_type')

        furniture_size = request.POST.get('furniture_size')

        furniture_thickness = request.POST.get('furniture_thickness')

        furniture_material = request.POST.get('furniture_material')

        furniture_features = request.POST.get('furniture_features')

        

        # Appliance Details

        appliance_type = request.POST.get('appliance_type')

        appliance_power_rating = request.POST.get('appliance_power_rating')

        appliance_voltage = request.POST.get('appliance_voltage')

        appliance_warranty = request.POST.get('appliance_warranty')



        # Validation

        errors = []

        if not product_name or len(product_name.strip()) == 0:

            errors.append("Product name is required")

        if not product_description or len(product_description.strip()) == 0:

            errors.append("Description is required")

        

        # At least one pricing option

        if not rent_price_per_day and not rent_price_per_month:

            errors.append("Enter at least one rental price (daily or monthly)")



        # Validate prices

        try:

            if rent_price_per_day:

                rent_price_per_day = float(rent_price_per_day)

                if rent_price_per_day < 0:

                    errors.append("Daily price cannot be negative")

        except ValueError:

            errors.append("Daily price must be a valid number")



        try:

            if rent_price_per_month:

                rent_price_per_month = float(rent_price_per_month)

                if rent_price_per_month < 0:

                    errors.append("Monthly price cannot be negative")

        except ValueError:

            errors.append("Monthly price must be a valid number")



        # Validate security deposit

        try:

            if security_deposit:

                security_deposit = float(security_deposit)

                if security_deposit < 0:

                    errors.append("Security deposit cannot be negative")

            else:

                security_deposit = 0

        except ValueError:

            errors.append("Security deposit must be a valid number")



        # Validate quantity

        try:

            stock_quantity = int(stock_quantity) if stock_quantity else 1

            if stock_quantity <= 0:

                errors.append("Quantity must be at least 1")

        except ValueError:

            errors.append("Quantity must be a valid number")



        # Validate image

        if product_image:

            allowed_types = ['image/jpeg', 'image/png', 'image/webp']

            if product_image.content_type not in allowed_types:

                errors.append("Product image must be JPEG, PNG, or WebP")



        if errors:

            categories = tbl_categorys.objects.all()

            for error in errors:

                messages.error(request, error)

            return render(request, 'Vendor/edit_product.html', {

                'product': product,

                'categories': categories,

                'form_data': request.POST

            })



        try:

            # Update product

            product.product_name = product_name.strip()

            product.product_description = product_description.strip()

            product.brand = brand.strip() if brand else None

            product.rent_price_per_day = rent_price_per_day if rent_price_per_day else None

            product.rent_price_per_month = rent_price_per_month if rent_price_per_month else None

            product.security_deposit = security_deposit

            product.stock_quantity = stock_quantity

            product.is_available = is_available

            

            if product_image:

                product.product_image = product_image

            

            product.save()



            # Update category-specific details

            category = product.category

            if category.category_name.lower() == 'furniture':

                furniture_details, created = tbl_furniture_details.objects.get_or_create(product=product)

                furniture_details.furniture_type = furniture_type.strip()

                furniture_details.size = furniture_size.strip()

                furniture_details.thickness = furniture_thickness.strip() if furniture_thickness else ''

                furniture_details.material = furniture_material.strip()

                furniture_details.extra_features = furniture_features.strip() if furniture_features else None

                furniture_details.save()

            

            elif category.category_name.lower() == 'appliances':

                appliance_details, created = tbl_appliance_details.objects.get_or_create(product=product)

                appliance_details.operation_type = appliance_type.strip()

                appliance_details.power_rating = appliance_power_rating.strip()

                appliance_details.voltage = appliance_voltage.strip()

                appliance_details.warranty_period = appliance_warranty.strip() if appliance_warranty else ''

                appliance_details.save()



            messages.success(request, f"âœ… Product '{product_name}' updated successfully!")

            return redirect('vendor_products')



        except Exception as e:

            import traceback

            error_details = traceback.format_exc()

            print(f"Error updating product: {error_details}")

            messages.error(request, f"Error updating product: {str(e)}")

            return redirect('edit_product', product_id=product_id)



    categories = tbl_categorys.objects.all()

    

    # Get category-specific details

    furniture_details = None

    appliance_details = None

    

    if hasattr(product, 'furniture_details'):

        furniture_details = product.furniture_details

    if hasattr(product, 'appliance_details'):

        appliance_details = product.appliance_details

    

    return render(request, 'Vendor/edit_product.html', {

        'product': product,

        'categories': categories,

        'furniture_details': furniture_details,

        'appliance_details': appliance_details

    })





def delete_product(request, product_id):

    """Delete a product"""

    try:

        vendor = tbl_vendor.objects.get(login_id=request.session.get('vendor'))

        product = tbl_product.objects.get(product_id=product_id, vendor=vendor)

        

        product_name = product.product_name

        product.delete()

        

        messages.success(request, f"âœ… Product '{product_name}' deleted successfully!")

        return redirect('vendor_products')

        

    except tbl_vendor.DoesNotExist:

        messages.error(request, "Vendor account not found")

        return redirect('vendor_home')

    except tbl_product.DoesNotExist:

        messages.error(request, "Product not found or you don't have permission to delete it")

        return redirect('vendor_products')

    except Exception as e:

        messages.error(request, f"Error deleting product: {str(e)}")

        return redirect('vendor_products')





def get_subcategories(request):

    """AJAX endpoint to get subcategories by category"""

    category_id = request.GET.get('category_id')

    

    if not category_id:

        return JsonResponse({'subcategories': []})

    

    try:

        subcategories = tbl_subcategory.objects.filter(

            category_id=category_id

        ).values('subcategory_id', 'subcategory_name')

        return JsonResponse({

            'subcategories': list(subcategories)

        })

    except Exception as e:

        return JsonResponse({'error': str(e)}, status=400)





# ======================== ORDER MANAGEMENT ========================



def vendor_orders(request):

    """Display orders for vendor's products"""

    if 'vendor' not in request.session:

        messages.warning(request, "Please login to view orders")

        return redirect('login')

    

    try:

        vendor = tbl_vendor.objects.get(login_id=request.session['vendor'])

        

        # Get all orders containing products from this vendor

        vendor_orders = tbl_order.objects.filter(

            items__product__vendor=vendor

        ).distinct().order_by('-created_at')

        

        # Get filter options

        status_filter = request.GET.get('status')

        if status_filter:

            vendor_orders = vendor_orders.filter(status=status_filter)

        
        # Add contextual info for each order
        orders_with_info = []
        for order in vendor_orders:
            order.items_count = order.items.count()
            orders_with_info.append(order)

        context = {

            'vendor': vendor,

            'orders': orders_with_info,

            'status_choices': tbl_order.ORDER_STATUS_CHOICES,

        }

        return render(request, 'Vendor/vendor_orders.html', context)

    

    except tbl_vendor.DoesNotExist:

        messages.error(request, "Vendor account not found")

        return redirect('vendor_home')





def order_details(request, order_id):

    """Display order details with vendor's products only"""

    if 'vendor' not in request.session:

        messages.warning(request, "Please login to view order details")

        return redirect('login')

    

    try:

        vendor = tbl_vendor.objects.get(login_id=request.session['vendor'])

        order = tbl_order.objects.get(order_id=order_id)

        

        # Get only items from this vendor

        vendor_items = order.items.filter(product__vendor=vendor)

        

        # Check if this vendor has any items in this order

        if not vendor_items.exists():

            messages.error(request, "You don't have access to this order")

            return redirect('vendor_orders')

        

        context = {

            'order': order,

            'vendor_items': vendor_items,

            'vendor': vendor,

        }

        

        return render(request, 'Vendor/order_details.html', context)

    

    except (tbl_vendor.DoesNotExist, tbl_order.DoesNotExist):

        messages.error(request, "Order not found")

        return redirect('vendor_orders')





@require_http_methods(["POST"])

def accept_order(request, order_id):

    """Accept an order - mark as confirmed"""

    if 'vendor' not in request.session:

        messages.warning(request, "Please login to accept orders")

        return redirect('login')

    

    try:

        vendor = tbl_vendor.objects.get(login_id=request.session['vendor'])

        order = tbl_order.objects.get(order_id=order_id)

        

        # Check if vendor has items in this order

        vendor_items = order.items.filter(product__vendor=vendor)

        if not vendor_items.exists():

            messages.error(request, "You don't have access to this order")

            return redirect('vendor_orders')

        

        # Only allow accepting pending orders

        if order.status != 'pending':

            messages.warning(request, "Only pending orders can be accepted")

            return redirect('order_details', order_id=order_id)

        

        # Update order status
        order.status = 'confirmed'
        order.save()
        
        # Send professional HTML email to customer
        try:
            from django.template.loader import render_to_string
            from django.urls import reverse
            
            subject = f"Order Confirmed - {order.order_number}"
            content = f"Hello {order.customer.name},\n\nGood news! Your order #{order.order_number} has been accepted by the vendor. Please proceed to payment to finalize your booking.\n\nThank you for choosing UrbanRentals."
            
            # Context for HTML email
            email_context = {
                'order': order,
                'order_items': order.items.all(),
                'dashboard_url': request.build_absolute_uri(reverse('customer_home')),
            }
            
            html_message = render_to_string('Vendor/email_order_confirmed.html', email_context)
            
            if send_email_notification(subject, content, order.customer.email, html_body=html_message):
                messages.success(request, f"Order {order.order_number} accepted successfully! Email sent to customer.")
            else:
                messages.success(request, f"Order {order.order_number} accepted successfully! (Email failed)")
        except Exception as e:
            print(f"Error sending confirmation email: {e}")
            messages.success(request, f"Order {order.order_number} accepted successfully!")
        
    except (tbl_vendor.DoesNotExist, tbl_order.DoesNotExist):

        messages.error(request, "Order not found")

    except Exception as e:

        messages.error(request, f"Error accepting order: {str(e)}")

    

    return redirect('order_details', order_id=order_id)





@require_http_methods(["POST"])

def reject_order(request, order_id):

    """Reject an order - mark as cancelled"""

    if 'vendor' not in request.session:

        messages.warning(request, "Please login to reject orders")

        return redirect('login')

    

    try:

        vendor = tbl_vendor.objects.get(login_id=request.session['vendor'])

        order = tbl_order.objects.get(order_id=order_id)

        

        # Check if vendor has items in this order

        vendor_items = order.items.filter(product__vendor=vendor)

        if not vendor_items.exists():

            messages.error(request, "You don't have access to this order")

            return redirect('vendor_orders')

        

        # Only allow rejecting pending orders

        if order.status != 'pending':

            messages.warning(request, "Only pending orders can be rejected")

            return redirect('order_details', order_id=order_id)

        

        # Update order status
        order.status = 'cancelled'
        order.save()
        
        # Send email to customer
        subject = f"Order Cancelled - {order.order_number}"
        content = f"Hello {order.customer.name},\n\nWe regret to inform you that your order #{order.order_number} has been declined by the vendor. Please check with other vendors.\n\nThank you for choosing UrbanRentals."
        
        if send_email_notification(subject, content, order.customer.email):
            messages.success(request, f"Order {order.order_number} rejected successfully! Email sent to customer.")
        else:
            messages.success(request, f"Order {order.order_number} rejected successfully! (Email failed)")
        
    except (tbl_vendor.DoesNotExist, tbl_order.DoesNotExist):

        messages.error(request, "Order not found")

    except Exception as e:

        messages.error(request, f"Error rejecting order: {str(e)}")

    

    return redirect('vendor_orders')


@require_http_methods(["POST"])
def mark_delivered(request, order_id):
    """Mark order as delivered after payment"""
    if 'vendor' not in request.session:
        messages.warning(request, "Please login first")
        return redirect('login') 
        
    try:
        vendor = tbl_vendor.objects.get(login_id=request.session['vendor'])
        order = tbl_order.objects.get(order_id=order_id)
        
        # Check if order belongs to vendor
        if not order.items.filter(product__vendor=vendor).exists():
            messages.error(request, "Access denied")
            return redirect('vendor_orders')
            
        if order.status != 'paid':
            messages.error(request, "Order must be paid before marking as delivered")
            return redirect('order_details', order_id=order_id)
            
        order.status = 'delivered'
        order.save()
        
        # Send professional HTML email to customer
        try:
            from django.template.loader import render_to_string
            from django.urls import reverse
            
            subject = f"Order Delivered - {order.order_number}"
            content = f"Hello {order.customer.name},\n\nYour order has been delivered successfully. Please confirm if you received the items.\n\nThank you for choosing UrbanRentals!"
            
            # Context for HTML email
            email_context = {
                'order': order,
                'order_items': order.items.all(),
                'track_url': request.build_absolute_uri(reverse('my_orders')),
                'review_url': request.build_absolute_uri(reverse('my_orders')),
                'browse_url': request.build_absolute_uri(reverse('browse_products')),
            }
            
            html_message = render_to_string('Vendor/email_order_delivered.html', email_context)
            
            if send_email_notification(subject, content, order.customer.email, html_body=html_message):
                messages.success(request, f"Order {order.order_number} marked as delivered! Customer notified with a beautiful email.")
            else:
                messages.success(request, f"Order {order.order_number} marked as delivered! (Email failed)")
        except Exception as e:
            print(f"Error sending delivery email: {e}")
            messages.success(request, f"Order {order.order_number} marked as delivered!")
             
    except (tbl_vendor.DoesNotExist, tbl_order.DoesNotExist):
        messages.error(request, "Order not found")
        
    return redirect('order_details', order_id=order_id)


# ======================== RETURN & PICKUP MANAGEMENT ========================



def vendor_return_requests(request):

    """Vendor views all return requests for their products"""

    try:

        vendor = tbl_vendor.objects.get(login_id=request.session.get('vendor'))

    except tbl_vendor.DoesNotExist:

        messages.error(request, "Vendor not found")

        return redirect('vendor_home')

    

    try:

        # Get return requests for this vendor's products

        return_requests = tbl_return_request.objects.filter(

            order_item__product__vendor=vendor

        ).select_related('customer', 'order_item__product', 'order').order_by('-request_date')

        

        # Separate by status

        pending = return_requests.filter(status='requested')

        scheduled = return_requests.filter(status='scheduled')

        picked_up = return_requests.filter(status='picked_up')

        completed = return_requests.filter(status='completed')

        

        context = {

            'all_returns': return_requests,

            'pending': pending,

            'scheduled': scheduled,

            'picked_up': picked_up,

            'completed': completed,

            'page_title': 'Return & Pickup Requests'

        }

        return render(request, 'Vendor/return_requests.html', context)

    except Exception as e:

        import traceback

        print(f"Error loading returns: {str(e)}")

        print(traceback.format_exc())

        messages.error(request, f"Error loading returns: {str(e)}")

        # return redirect('vendor_home')

        raise e





def schedule_pickup(request, return_id):

    """Vendor schedules pickup for a return request"""

    try:

        vendor = tbl_vendor.objects.get(login_id=request.session.get('vendor'))

        return_req = tbl_return_request.objects.get(return_id=return_id)

        

        # Verify this return is for vendor's product

        if return_req.order_item.product.vendor != vendor:

            messages.error(request, 'Unauthorized access to return request.')
            return redirect('vendor_return_requests')



        if request.method == 'POST':

            pickup_date = request.POST.get('pickup_scheduled_date')



            if not pickup_date:

                messages.error(request, 'Please select a date.')
                return redirect('vendor_return_requests')



            # Update return request

            return_req.pickup_scheduled_date = pickup_date

            return_req.status = 'scheduled'

            return_req.save()

            # Send email to customer
            try:
                from datetime import datetime
                try:
                    pickup_date_obj = datetime.strptime(pickup_date, '%Y-%m-%d').date()
                    formatted_date = pickup_date_obj.strftime('%B %d, %Y')
                except ValueError:
                    formatted_date = pickup_date

                customer_email = return_req.customer.email
                customer_name = return_req.customer.name
                product_name = return_req.order_item.product.product_name

                subject = f"Return Pickup Scheduled - {product_name}"
                body = f"""Hello {customer_name},

Your return request for '{product_name}' has been processed.

Our pickup team is scheduled to visit your location on: {formatted_date}

Please ensure the item is ready for inspection and handover.

Thank you for choosing Urban Rentals.
"""
                send_email_notification(subject, body, customer_email)
            except Exception as mail_error:
                print(f"Error sending mail: {mail_error}")

            messages.success(request, f'Pickup scheduled successfully for {pickup_date} and notification sent.')
            return redirect('vendor_return_requests')

    except tbl_return_request.DoesNotExist:

        messages.error(request, 'Return request not found.')
        return redirect('vendor_return_requests')

    except Exception as e:

        messages.error(request, f'Error: {str(e)}')
        return redirect('vendor_return_requests')



    return redirect('vendor_return_requests')

def complete_pickup(request, return_id):

    """Vendor completes pickup and assesses defects"""

    try:

        vendor = tbl_vendor.objects.get(login_id=request.session.get('vendor'))

        return_req = tbl_return_request.objects.get(return_id=return_id)

        

        # Verify authorization

        if return_req.order_item.product.vendor != vendor:

            messages.error(request, "Unauthorized")

            return redirect('vendor_home')

        

        if request.method == 'POST':

            # Defect assessment

            defect_status = request.POST.get('defect_status', 'no_defect')

            defect_description = request.POST.get('defect_description', '')

            defect_image = request.FILES.get('defect_image')

            actual_return_date = request.POST.get('actual_return_date')

            

            # Update return request

            return_req.defect_status = defect_status

            return_req.defect_description = defect_description if defect_description else None

            if defect_image:

                return_req.defect_image = defect_image

            

            # Set actual return date

            if actual_return_date:
                try:
                    if 'T' in actual_return_date:
                        return_req.actual_return_date = datetime.strptime(actual_return_date, '%Y-%m-%dT%H:%M').date()
                    elif len(actual_return_date) == 16 and ':' in actual_return_date:
                        return_req.actual_return_date = datetime.strptime(actual_return_date, '%Y-%m-%d %H:%M').date()
                    else:
                        return_req.actual_return_date = datetime.strptime(actual_return_date, '%Y-%m-%d').date()
                except ValueError:
                    return_req.actual_return_date = datetime.strptime(actual_return_date[:10], '%Y-%m-%d').date()

            

            # Calculate deductions

            return_req.calculate_deductions(extra_day_percentage=5)

            

            # Mark as picked up

            return_req.pickup_completed_date = datetime.now()

            return_req.status = 'completed'

            return_req.save()
            
            # ========== RESTORE PRODUCT STOCK AFTER RETURN COMPLETION ==========
            from UrbanRentalsProject.utils import restore_product_stock
            
            success, message, product = restore_product_stock(return_req.order_item)
            if success:
                print(f"[SUCCESS] {message}")
                messages.success(request, f"Return {return_id} completed. Refund: ₹{return_req.refund_amount}. {message}")
            else:
                print(f"[WARNING] Stock restoration failed: {message}")
                messages.success(request, f"Return {return_id} completed. Refund: ₹{return_req.refund_amount}. (Stock restoration warning: {message})")
            # ========================================================================

            return redirect('vendor_return_requests')

        

        context = {
            'return_req': return_req,
            'order_item': return_req.order_item,
            'product': return_req.order_item.product,
            'customer': return_req.customer,
            'page_title': 'Complete Pickup & Assessment'
        }

        return render(request, 'Vendor/complete_pickup.html', context)

    except tbl_return_request.DoesNotExist:

        messages.error(request, "Return request not found")

        return redirect('vendor_return_requests')

    except Exception as e:

        messages.error(request, f"Error: {str(e)}")

        return redirect('vendor_return_requests')


def edit_profile(request):
    vendor_login_id = request.session.get('vendor')
    if not vendor_login_id:
        return redirect('login') 

    try:
        vendor = tbl_vendor.objects.get(login_id=vendor_login_id)
        
        if request.method == 'POST':
            name = request.POST.get('name')
            email = request.POST.get('email')
            phone = request.POST.get('phone')
            address = request.POST.get('address')
            gstin = request.POST.get('gstin')
            
            # Basic update logic
            vendor.name = name
            vendor.email = email
            vendor.phone = phone
            vendor.address = address
            vendor.gstin = gstin
            
            try:
                vendor.save()
                messages.success(request, 'Profile updated successfully')
                return redirect('vendor_profile')
            except Exception as e:
                messages.error(request, f'Error updating profile: {str(e)}')
        
        context = {
            'vendor': vendor
        }
        return render(request, 'Vendor/edit_profile.html', context)
        
    except tbl_vendor.DoesNotExist:
        messages.error(request, 'Vendor details not found')
        return redirect('vendor_home')


def vendor_report(request):
    """
    Date-wise sales report for vendor (Order Items).
    """
    if 'vendor' not in request.session:
        return redirect('login')
    
    try:
        vendor = tbl_vendor.objects.get(login_id=request.session['vendor'])
        
        # Filter order items related to this vendor
        order_items = tbl_order_item.objects.filter(product__vendor=vendor).order_by('-order__created_at')
        
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')

        if from_date and to_date:
            order_items = order_items.filter(order__created_at__date__range=[from_date, to_date])
        elif from_date:
            order_items = order_items.filter(order__created_at__date__gte=from_date)
        elif to_date:
            order_items = order_items.filter(order__created_at__date__lte=to_date)

        return render(request, 'Vendor/vendor_report.html', {
            'order_items': order_items,
            'from_date': from_date,
            'to_date': to_date
        })

    except tbl_vendor.DoesNotExist:
        messages.error(request, 'Vendor details not found')
        return redirect('vendor_home')


def vendor_report_export(request):
    """
    Export date-wise sales report to Excel for vendor.
    """
    if 'vendor' not in request.session:
        return redirect('login')
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        vendor = tbl_vendor.objects.get(login_id=request.session['vendor'])
        
        order_items = tbl_order_item.objects.filter(product__vendor=vendor).order_by('-order__created_at')
        
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')

        if from_date and to_date:
            order_items = order_items.filter(order__created_at__date__range=[from_date, to_date])
        elif from_date:
            order_items = order_items.filter(order__created_at__date__gte=from_date)
        elif to_date:
            order_items = order_items.filter(order__created_at__date__lte=to_date)

        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Report"

        # Headers
        headers = ['Order #', 'Product', 'Customer', 'Date', 'Status', 'Rent Type', 'Duration', 'Total Rent', 'Deposit', 'Grand Total']
        ws.append(headers)
        
        # Style header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid") # Using existing green theme
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for item in order_items:
            try:
                # Handle potentially missing attributes gracefully
                duration = getattr(item, 'rental_duration', 'N/A')
                deposit = getattr(item, 'security_deposit', 0)
                grand_total = getattr(item, 'grand_total', 0)
                
                row = [
                    item.order.order_number,
                    item.product.product_name,
                    item.order.customer.name,
                    item.order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    item.order.get_status_display(),
                    item.get_rent_type_display(),
                    duration,
                    item.total_rent,
                    deposit,
                    grand_total
                ]
                ws.append(row)
            except Exception as e:
                # Fallback if attribute access fails
                print(f"Error processing item {item.id}: {e}")
                continue

        # Auto-adjust column width
        for column_cells in ws.columns:
            length = max(len(str(cell.value) or "") for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="my_sales_report.xlsx"'
        wb.save(response)
        
        return response

    except tbl_vendor.DoesNotExist:
        messages.error(request, 'Vendor details not found')
        return redirect('vendor_home')
    except Exception as e:
        messages.error(request, f"Error generating Excel report: {str(e)}")
        # Fallback to CSV if Excel fails? Or just redirect
        return redirect('vendor_report')


