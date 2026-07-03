from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from AdminApp.models import *
from GuestApp.models import tbl_customer, tbl_vendor
from CustomerApp.models import tbl_order
from django.contrib import messages
from django.shortcuts import get_object_or_404
from django.views.decorators.http import require_http_methods
import json
from .decorators import admin_required
import smtplib
from email.message import EmailMessage
from django.template.loader import render_to_string

def send_email_notification(subject, body, to_email, html_body=None):
    """
    Helper function to send email notifications.
    Replace 'your mailid' and '2 step verification password' with actual credentials.
    """
    try:
        msg = EmailMessage()
        msg.set_content(body)
        if html_body:
            msg.add_alternative(html_body, subtype='html')

        msg['Subject'] = subject
        msg['From'] = "urbanrentalsofficial@gmail.com"  # Create a dedicated email for your app
        msg['To'] = to_email
        
        # Connect to Gmail's SMTP server
        server = smtplib.SMTP_SSL("smtp.gmail.com", 465)
        server.login("urbanrentalsofficial@gmail.com", "kilw hwzc jdkz xlmp")
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print(f"Error sending email: {e}")
        return False


from django.db.models import Sum, Count, F
from django.db.models.functions import TruncMonth, TruncYear
import json
from django.views.decorators.cache import cache_control

# Create your views here.
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@admin_required
def admin_home(request):
    # Dashboard Statistics
    STATUS_CHOICES_FOR_REVENUE = ['paid', 'completed', 'rented', 'returned', 'confirmed']
    
    total_customers = tbl_customer.objects.count()
    total_vendors = tbl_vendor.objects.count()
    total_orders = tbl_order.objects.count()
    total_revenue = tbl_order.objects.filter(status__in=STATUS_CHOICES_FOR_REVENUE).aggregate(
        total=Sum('grand_total')
    )['total'] or 0

    # Recent Orders
    recent_orders = tbl_order.objects.select_related('customer').order_by('-created_at')[:5]

    # Chart Data: Monthly Sales for Current Year
    from datetime import datetime
    import calendar
    current_year = datetime.now().year

    # Filter for current year
    monthly_sales = tbl_order.objects.filter(
        status__in=STATUS_CHOICES_FOR_REVENUE,
        created_at__year=current_year
    ).annotate(month=TruncMonth('created_at'))\
        .values('month')\
        .annotate(total=Sum('grand_total'))\
        .order_by('month')
    
    monthly_sales_list = list(monthly_sales) 
    
    # Fill in all 12 months
    sales_data_dict = {}
    for entry in monthly_sales_list:
        month_val = entry.get('month')
        total_val = entry.get('total')
        
        if month_val:
            try:
                # Handle cases where month might be a datetime object or a string (SQLite oddity)
                mon_num = month_val.month if hasattr(month_val, 'month') else int(str(month_val).split('-')[1])
                sales_data_dict[mon_num] = float(total_val or 0)
            except (ValueError, IndexError, AttributeError):
                continue
            
    sales_dates = []
    sales_amounts = []
    
    for i in range(1, 13):
        sales_dates.append(calendar.month_name[i])
        sales_amounts.append(sales_data_dict.get(i, 0.0))

    # Chart Data: Order Status Distribution
    order_status_counts = tbl_order.objects.values('status').annotate(count=Count('status'))
    status_labels = [entry['status'] for entry in order_status_counts]
    status_data = [entry['count'] for entry in order_status_counts]

    context = {
        'total_revenue': total_revenue,
        'total_bookings': total_orders, # Renamed for template consistency
        'today_bookings': tbl_order.objects.filter(created_at__date=datetime.now().date()).count(),
        'active_users': tbl_customer.objects.filter(login__status='active').count(),
        'active_vendors': tbl_vendor.objects.filter(login__status='active').count(),
        'sales_dates': sales_dates,     # Passing list directly for json_script
        'sales_amounts': sales_amounts,
        'status_labels': status_labels,
        'status_data': status_data,
        'orders': tbl_order.objects.all().order_by('-created_at')[:20],
    }
    return render(request, 'Admin/index.html', context)

@admin_required
def district(request):
    if request.method == 'POST':
        dis = request.POST.get('districtname').strip()

        # Field validation
        if not dis:
            messages.error(request, "District name cannot be empty!")
            return redirect('district')

        # Check duplicate (case-insensitive)
        if tbl_district.objects.filter(district_name__iexact=dis).exists():
            messages.error(request, "District already exists!")
            return redirect('district')

        # Save new district
        tbl_district.objects.create(district_name=dis)
        messages.success(request, "District added successfully!")
        return redirect('district')

    return render(request, 'admin/district.html')

@admin_required
def viewdistrict(request):
    dis = tbl_district.objects.all()
    return render(request, 'admin/viewdistrict.html',{'d': dis})

@admin_required
def deletedistrict(request, did):
    dis = get_object_or_404(tbl_district, district_id=did)

    dis.delete()
    messages.success(request, "District deleted successfully!")

    return redirect('viewdistrict')

@admin_required
def editdistrict(request, did):
    dis = get_object_or_404(tbl_district, district_id=did)
    district_list = tbl_district.objects.all()  # For duplicate check

    if request.method == 'POST':
        disname = request.POST.get('districtname').strip()

        # Empty field check
        if not disname:
            messages.error(request, "District name cannot be empty!")
            return redirect('editdistrict', did=did)

        # Already exists check (except current)
        if tbl_district.objects.filter(district_name__iexact=disname).exclude(district_id=did).exists():
            messages.error(request, "District already exists!")
            return redirect('editdistrict', did=did)

        # Save update
        dis.district_name = disname
        dis.save()

        messages.success(request, "District updated successfully!")
        return redirect('viewdistrict')

    return render(request, 'admin/editdistrict.html', {'d': dis})



@admin_required
def location(request):
    # Load all districts for dropdown
    districtdata = tbl_district.objects.all().order_by('district_name')

    if request.method == 'POST':
        loc_name = request.POST.get('location_name', '').strip()
        district_id = request.POST.get('district_id')

        # Validation – empty fields
        if not loc_name or not district_id:
            messages.error(request, "All fields are required!")
            return redirect('location')

        # Check if location already exists in same district
        if tbl_location.objects.filter(
            location_name__iexact=loc_name,
            district_id=district_id
        ).exists():
            messages.error(request, "Location already exists!")
            return redirect('location')

        # Save location
        district_obj = tbl_district.objects.get(district_id=district_id)
        tbl_location.objects.create(
            location_name=loc_name,
            district_id=district_obj
        )

        messages.success(request, "Location added successfully!")
        return redirect('location')

    return render(request, 'admin/location.html', {'districtdata': districtdata})

@admin_required
def viewlocation(request):
    locationdata = tbl_location.objects.select_related('district_id').all()
    return render(request, 'admin/viewlocation.html', {'locationdata': locationdata})

@admin_required
def editlocation(request, lid):
    loc = tbl_location.objects.get(location_id=lid)
    districtdata = tbl_district.objects.all()

    if request.method == 'POST':
        loc.location_name = request.POST.get('location_name')
        district_id = request.POST.get('district_id')

        # Update district
        loc.district_id = tbl_district.objects.get(district_id=district_id)

        # Save changes
        loc.save()

        messages.success(request, "Location updated successfully!")
        return redirect('viewlocation')

    return render(request, 'admin/editlocation.html', {
        'loc': loc,
        'districtdata': districtdata
    })

@admin_required
def deletelocation(request, lid):
    # Safely fetch the location or show 404 if not found
    loc = get_object_or_404(tbl_location, location_id=lid)

    # Delete the record
    loc.delete()

    # Success message
    messages.success(request, "Location deleted successfully!")

    # Redirect back to the location table
    return redirect('viewlocation')

@admin_required
def category(request):
    if request.method == "POST":
        category_name = request.POST.get("category_name")
        description = request.POST.get("description")
        category_image = request.FILES.get("category_image")

        # Validation
        if not category_name:
            messages.error(request, "Category name is required")
            return redirect("categoryreg")

        # Prevent duplicate category
        if tbl_categorys.objects.filter(category_name__iexact=category_name).exists():
            messages.warning(request, "Category already exists")
            return redirect("categoryreg")

        # Save category
        tbl_categorys.objects.create(
            category_name=category_name,
            description=description,
            category_image=category_image
        )

        messages.success(request, "Category added successfully")
        return redirect("categoryreg")

    return render(request, "Admin/categoryreg.html")

@admin_required
def viewcategory(request):
    categorydata = tbl_categorys.objects.all().order_by('category_name')
    return render(request, 'Admin/viewcategory.html', {
        'categorydata': categorydata
    })

@admin_required
def editcategory(request, id):
    category = get_object_or_404(tbl_categorys, category_id=id)

    if request.method == 'POST':
        category.category_name = request.POST.get('category_name')
        category.description = request.POST.get('category_description')

        # Image update (only if new image selected)
        if 'category_image' in request.FILES:
            category.category_image = request.FILES['category_image']

        category.save()
        messages.success(request, "Category updated successfully")
        return redirect('viewcategory')

    return render(request, 'Admin/editcategory.html', {'category': category})


@admin_required
def deletecategory(request, id):
    category = get_object_or_404(tbl_categorys, category_id=id)

    if request.method == 'POST':
        category.delete()
        messages.success(request, "Category deleted successfully.")

    return redirect('viewcategory')


@admin_required
def customer_verification(request):
    """Display all customers for verification"""
    customers = tbl_customer.objects.all().order_by('-customer_id')
    
    # Calculate customer statistics
    verified_count = customers.filter(is_verified=True).count()
    rejected_count = customers.filter(is_rejected=True).count()
    pending_count = customers.exclude(is_verified=True).exclude(is_rejected=True).count()
    
    return render(request, 'Admin/customerverification.html', {
        'customers': customers,
        'verified_count': verified_count,
        'rejected_count': rejected_count,
        'pending_count': pending_count,
    })


@admin_required
def verify_customer(request, customer_id):
    """Verify a customer and enable login"""
    customer = get_object_or_404(tbl_customer, customer_id=customer_id)
    customer.is_verified = True
    customer.is_rejected = False
    customer.status = "Verified"
    customer.save()
    
    # Also update the associated login status to Active (if needed)
    if customer.login:
        customer.login.status = "Active"
        customer.login.save()
    
    # Send verification email
    from django.urls import reverse
    
    subject = "Account Verified - UrbanRentals"
    body = f"Hello {customer.name},\n\nYour account has been successfully verified by our admin. You can now login to UrbanRentals."
    
    html_body = render_to_string('Admin/email_customer_verified.html', {
        'customer': customer,
        'login_url': request.build_absolute_uri(reverse('login'))
    })
    
    if send_email_notification(subject, body, customer.email, html_body=html_body):
        messages.success(request, f"✅ Customer '{customer.name}' has been VERIFIED! Email notification sent.")
    else:
        messages.warning(request, f"✅ Customer '{customer.name}' verified, but email notification failed. Check server logs.")

    return redirect('customer_verification')


@admin_required
def reject_customer(request, customer_id):
    """Reject a customer and prevent login"""
    customer = get_object_or_404(tbl_customer, customer_id=customer_id)
    customer.is_rejected = True
    customer.is_verified = False
    customer.status = "Rejected"
    customer.save()
    
    # Send rejection notification
    subject = "Account Application Update - UrbanRentals"
    body = f"Hello {customer.name},\n\nWe regret to inform you that your registration request for UrbanRentals has been declined. Please contact support for more details."

    if send_email_notification(subject, body, customer.email):
        messages.error(request, f"❌ Customer '{customer.name}' has been REJECTED! Email notification sent.")
    else:
        messages.warning(request, f"❌ Customer '{customer.name}' rejected, but email notification failed.")

    return redirect('customer_verification')


@admin_required
def subcategory_registration(request):
    categorydata = tbl_categorys.objects.all().order_by('category_name')

    if request.method == "POST":
        category_id = request.POST.get('category_id')
        subcategory_name = request.POST.get('subcategory_name')
        description = request.POST.get('description')
        subcategory_image = request.FILES.get('subcategory_image')

        # Validation
        if not category_id:
            messages.error(request, "Please select a category")
            return redirect('subcategory_registration')

        if not subcategory_name:
            messages.error(request, "Subcategory name is required")
            return redirect('subcategory_registration')

        # Duplicate check (same subcategory under same category)
        if tbl_subcategory.objects.filter(
            category_id=category_id,
            subcategory_name__iexact=subcategory_name
        ).exists():
            messages.error(request, "This subcategory already exists under selected category")
            return redirect('subcategory_registration')

        # Image validation (optional)
        if subcategory_image:
            if subcategory_image.size > 2 * 1024 * 1024:
                messages.error(request, "Image size must be under 2MB")
                return redirect('subcategory_registration')

            if not subcategory_image.content_type.startswith('image'):
                messages.error(request, "Only image files are allowed")
                return redirect('subcategory_registration')

        # Save subcategory
        tbl_subcategory.objects.create(
            category_id=category_id,
            subcategory_name=subcategory_name,
            description=description,
            subcategory_image=subcategory_image
        )

        messages.success(request, "Subcategory registered successfully")
        return redirect('subcategory_registration')

    return render(request, 'Admin/subcategoryreg.html', {
        'categorydata': categorydata
    })

@admin_required
def viewsubcategory(request):
    # Fetch all subcategories
    subcategories = tbl_subcategory.objects.all().order_by('subcategory_name')

    context = {
        'subcategories': subcategories
    }

    return render(request, 'Admin/viewsubcategory.html', context)


@admin_required
def editsubcategory(request, id):
    # Fetch subcategory safely
    subcategory = get_object_or_404(tbl_subcategory, pk=id)

    # Fetch all categories for dropdown
    categories = tbl_categorys.objects.all().order_by('category_name')

    if request.method == "POST":
        subcategory_name = request.POST.get('subcategory_name')
        category_id = request.POST.get('category_id')
        description = request.POST.get('description')
        subcategory_image = request.FILES.get('subcategory_image')

        # Update fields
        subcategory.subcategory_name = subcategory_name
        subcategory.category_id = category_id
        subcategory.description = description
        
        if subcategory_image:
            subcategory.subcategory_image = subcategory_image
            
        subcategory.save()
        messages.success(request, "Subcategory updated successfully")
        return redirect('viewsubcategory')

    return render(request, 'Admin/editsubcategory.html', {
        'subcategory': subcategory,
        'categories': categories
    })

@admin_required
def deletesubcategory(request, id):
    subcategory = get_object_or_404(tbl_subcategory, pk=id)
    subcategory.delete()
    return redirect('viewsubcategory')


# ============ VENDOR VERIFICATION VIEWS ============

@admin_required
def vendor_verification(request):
    """Display vendor verification page with list of all vendors"""
    vendors = tbl_vendor.objects.select_related('location', 'location__district_id').all().order_by('-created_at')
    pending_count = vendors.filter(status='Pending').count()
    approved_count = vendors.filter(status='Approved').count()
    rejected_count = vendors.filter(status='Rejected').count()
    
    context = {
        'vendors': vendors,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count
    }
    
    return render(request, 'Admin/vendorverification.html', context)


@admin_required
def approve_vendor(request, vendor_id):
    """Approve a vendor"""
    vendor = get_object_or_404(tbl_vendor, vendor_id=vendor_id)
    vendor.status = 'Approved'
    vendor.save()
    
    # Also update the associated login status to Active
    if vendor.login:
        vendor.login.status = 'Active'
        vendor.login.save()
    
    # Send verification email
    from django.urls import reverse
    
    subject = "Vendor Account Approved - UrbanRentals"
    body = f"Hello {vendor.name},\n\nYour vendor account has been successfully verified and approved by our admin. You can now login to UrbanRentals and start listing your products."
    
    html_body = render_to_string('Admin/email_vendor_approved.html', {
        'vendor': vendor,
        'login_url': request.build_absolute_uri(reverse('login'))
    })
    
    if send_email_notification(subject, body, vendor.email, html_body=html_body):
        messages.success(request, f"✅ Vendor '{vendor.name}' has been APPROVED! Email notification sent.")
    else:
        messages.warning(request, f"✅ Vendor '{vendor.name}' approved, but email notification failed. Check server logs.")

    return redirect('vendor_verification')


@admin_required
def reject_vendor(request, vendor_id):
    """Reject a vendor"""
    vendor = get_object_or_404(tbl_vendor, vendor_id=vendor_id)
    vendor.status = 'Rejected'
    vendor.save()
    
    # Send rejection notification
    subject = "Vendor Application Update - UrbanRentals"
    body = f"Hello {vendor.name},\n\nWe regret to inform you that your vendor account application for UrbanRentals has been declined. Please ensure you have uploaded valid proofs and try again, or contact support."

    if send_email_notification(subject, body, vendor.email):
         messages.error(request, f"❌ Vendor '{vendor.name}' has been REJECTED! Email notification sent.")
    else:
        messages.warning(request, f"❌ Vendor '{vendor.name}' rejected, but email notification failed.")

    return redirect('vendor_verification')


import csv
from datetime import datetime

@admin_required
def admin_sales_report(request):
    """
    Date-wise sales report for admin (Orders).
    """
    orders = tbl_order.objects.all().order_by('-created_at')

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if from_date and to_date:
        orders = orders.filter(created_at__date__range=[from_date, to_date])
    elif from_date:
        orders = orders.filter(created_at__date__gte=from_date)
    elif to_date:
        orders = orders.filter(created_at__date__lte=to_date)

    return render(request, 'Admin/sales_report.html', {
        'orders': orders,
        'from_date': from_date,
        'to_date': to_date
    })


@admin_required
def admin_sales_report_export(request):
    """
    Export date-wise sales report to Excel.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    orders = tbl_order.objects.all().order_by('-created_at')

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if from_date and to_date:
        orders = orders.filter(created_at__date__range=[from_date, to_date])
    elif from_date:
        orders = orders.filter(created_at__date__gte=from_date)
    elif to_date:
        orders = orders.filter(created_at__date__lte=to_date)

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Headers
    headers = ['Order Number', 'Customer', 'Date', 'Status', 'Total Rent', 'Deposit', 'Grand Total']
    ws.append(headers)
    
    # Style header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid") # Green theme
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for order in orders:
        row = [
            order.order_number,
            order.customer.name,
            order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            order.get_status_display(),
            order.total_rent,
            order.total_security_deposit,
            order.grand_total
        ]
        ws.append(row)

    # Auto-adjust column width
    for column_cells in ws.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="admin_sales_report.xlsx"'
    wb.save(response)

    return response

